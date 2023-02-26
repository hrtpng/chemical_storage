from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from storage.models import Compound


class Command(BaseCommand):
    help = "Imports excel file to db."

    def add_arguments(self, parser):
        parser.add_argument("file_path", type=str)

    def handle(self, *args, **kwargs):
        file_path = kwargs["file_path"]

        excel = load_workbook(file_path)

        def get_data_from_strona(rack, page):
            row_index = 0
            for row in page.rows:
                row_index += 1
                if row_index > 1:
                    chem = []
                    for cell in row:
                        chem.append(cell.value)
                    Compound.objects.create(
                        formula=chem[1],
                        title=chem[2],
                        iupac=chem[3],
                        cas=chem[4],
                        tags=chem[5],
                        shelf=chem[6],
                        rack=rack,
                    ).save()

        for name in excel.sheetnames:
            rack = name.split(" ")[-1]
            page = excel[name]
            get_data_from_strona(rack, page)
        excel.close()

        self.stdout.write(self.style.SUCCESS("HI"))
